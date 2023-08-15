import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from settings import *


def save_file(data):
    # Saving the excell file
    data.save('data.xlsx')
    print('Data saved')


def col_to_name(col):
    col_num = col

    col_str = ''

    while col_num:
        # Set remainder from 1 .. 26
        remainder = col_num % 26

        if remainder == 0:
            remainder = 26

        # Convert the remainder to a character.
        col_letter = chr(ord('A') + remainder - 1)

        # Accumulate the column letters, right to left.
        col_str = col_letter + col_str

        # Get the next order of magnitude.
        col_num = int((col_num - 1) / 26)

    return col_str


def get_students_mails(sheet):
    result = []
    students_num = sheet.max_row - 1
    start_index = STUDENTS_MAILS_START_INDEX

    for i in range(start_index, students_num + start_index):
        result.append(sheet[f'B{i}'].value)  # Will not be affected if settings change
    
    return result


def get_subjects_names(sheet):
    result = []
    subjects_num = sheet.max_column - 2
    start_index = SUBJECTS_NAMES_START_INDEX

    for i in range(start_index, subjects_num + start_index):
        result.append(
            sheet[f'{col_to_name(i)}1'].value  # Will not be affected if settings change
        )
    
    return result


def send_mail(mails_list: list[str], subject: str, message_body: str):
    sender_mail = 'sender@gmail.com'  # Enter the sender email
    sender_app_password = 'app password'  # Turn 2-step Verification on and get your app password

    # Prepare smtp
    # Port 587 is for TLS connection
    server = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    server.starttls()

    # Login to the sending email
    server.login(sender_mail, sender_app_password)

    for receiver_mail in mails_list:
        # Preparing the message
        message = MIMEMultipart()
        message['Subject'] = subject

        # Attaching the body text
        message.attach(MIMEText(message_body, 'plain'))

        content = message.as_string()
        
        # Sending the mail
        server.sendmail(sender_mail, receiver_mail, content)
    
    # Terminate the session
    server.quit()
    
    print('Mail was sent to students')


def check(sheet):
    # Get mail addresses and subjects
    students_mails = get_students_mails(sheet)
    subjects_names = get_subjects_names(sheet)

    for i, student_mail in enumerate(students_mails):
        i += STUDENTS_MAILS_START_INDEX
        near_miss = []
        miss = []

        for j, subject_name in enumerate(subjects_names):
            j += SUBJECTS_NAMES_START_INDEX

            # Get the number of misses on this class
            number_miss = sheet[f'{col_to_name(j)}{i}'].value

            if number_miss == MAX_MISSED_CLASSES:
                near_miss.append(subject_name)
            elif number_miss > MAX_MISSED_CLASSES:
                miss.append(subject_name)
        
        message = ''
        if len(near_miss) > 0:
            message += "You can't miss these classes anymore or you will be kicked: {}\n".format(' '.join(near_miss))
        
        if len(miss) > 0:
            message += "You was kicked from these classes: {}\n".format(' '.join(miss))
        
        if len(message) == 0:
            message += "You are alright!"
        
        send_mail([student_mail], 'Class attendance report', message)


def main():
    # Load the excel file. It should be in the same folder as the 'main.py' file
    data = openpyxl.load_workbook('data.xlsx')

    # Get the sheet where the data is written
    sheet = data['Sheet1']

    run = True
    while run:
        request_type = int(input('Enter:\n \
                                 1 if ----> You want to add missing students\n \
                                 2 if ----> You want to send a mail with attendance report to students\n \
                                 3 if ----> You want to stop\n'))

        match request_type:
            case 1:
                subjects_names = get_subjects_names(sheet)

                for i, subject in enumerate(subjects_names):
                    i += SUBJECTS_NAMES_START_INDEX
                    miss_number = int(input(f'Enter the number of students who missed class {subject}: '))

                    if miss_number == 0:
                        continue
                    elif miss_number == 1:
                        student_id = int(input("Enter the student's id: "))
                        student_id += STUDENTS_MAILS_START_INDEX - 1

                        # Update value in sheet
                        sheet[f'{col_to_name(i)}{student_id}'].value += 1
                    else:
                        student_ids = list(map(int, 
                                               input("Enter the students' ids in one line with space between: ").split()))

                        for student_id in student_ids:
                            student_id += STUDENTS_MAILS_START_INDEX - 1

                            # Update value in sheet
                            sheet[f'{col_to_name(i)}{student_id}'].value += 1

                # Save the file
                save_file(data)
            case 2:
                # Check and send mails with warnings
                check(sheet)
            case 3:
                # End the program
                run = False
            case _:
                print("Unknown request, please type again!\n")


if __name__ == '__main__':
    main()
